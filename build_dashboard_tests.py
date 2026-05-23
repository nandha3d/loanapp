from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Palette ──────────────────────────────────────────────────────────────────
H_BG   = "1F3864"   # dark navy header
H_FG   = "FFFFFF"
S_ML   = "F4E6D3"   # micro-lending section - amber tint
S_AF   = "D3E8F4"   # auto finance - blue tint
S_CF   = "D3F4E0"   # chit funds - green tint
S_AG   = "F4F0D3"   # agent view - yellow tint
ML_HDR = "E67E22"
AF_HDR = "2980B9"
CF_HDR = "27AE60"
AG_HDR = "8E44AD"
PASS_BG = "C6EFCE"; FAIL_BG = "FFC7CE"; PEND_BG = "FFEB9C"; NA_BG = "E2EFDA"
WHITE  = "FFFFFF";  GRAY   = "F5F5F5"

thin  = Side(style="thin", color="CCCCCC")
bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)

def hfont(sz=10, bold=True, color=H_FG): return Font(name="Arial", bold=bold, size=sz, color=color)
def dfont(sz=9,  bold=False,color="000000"): return Font(name="Arial", bold=bold, size=sz, color=color)
def fill(c): return PatternFill("solid", fgColor=c)
def ctr(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def lft(): return Alignment(horizontal="left",   vertical="center", wrap_text=True)

def hdr_row(ws, r, cols, bg=H_BG):
    for c in range(1, cols+1):
        cell = ws.cell(r, c)
        cell.font = hfont(); cell.fill = fill(bg)
        cell.alignment = ctr(); cell.border = bdr
    ws.row_dimensions[r].height = 28

def sec_row(ws, r, label, cols, bg):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
    cell = ws.cell(r, 1, f"  {label}")
    cell.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    cell.fill = fill(bg); cell.alignment = lft(); cell.border = bdr
    ws.row_dimensions[r].height = 20

def data_row(ws, r, values, alt=False):
    bg = GRAY if alt else WHITE
    for c, v in enumerate(values, 1):
        cell = ws.cell(r, c, v)
        cell.font = dfont(); cell.fill = fill(bg)
        cell.alignment = lft(); cell.border = bdr
    ws.row_dimensions[r].height = 52


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — OVERVIEW
# ══════════════════════════════════════════════════════════════════════════════
ws0 = wb.active; ws0.title = "Overview"
ws0.sheet_view.showGridLines = False
ws0.column_dimensions["A"].width = 26
ws0.column_dimensions["B"].width = 72

ws0["A1"] = "LoanTrack — Dashboard KPI Automated Test Plan"
ws0["A1"].font = Font(name="Arial", bold=True, size=15, color=H_BG)
ws0["A2"] = "Three Modules × All Roles — Every KPI Verified Against DB"
ws0["A2"].font = Font(name="Arial", size=10, color="555555")

rows = [
    ("PURPOSE",
     "Each test case describes the exact DB state to seed, the precise SQL / API assertion to run, "
     "and the expected value. An automated agent runs every case in module-URL order and records "
     "Pass / Fail / Pending without fixing anything."),
    ("URL PATTERN",
     "/{module}/dashboard — e.g. /microlending/dashboard, /autofinance/dashboard, /chitfunds/dashboard"),
    ("MODULES COVERED",
     "Micro Lending (ML) • Auto Finance (AF) • Chit Funds (CF)"),
    ("ROLES COVERED",
     "superadmin • admin • agent (scoped view) — all tested against the same KPI assertions"),
    ("HOW TO RUN",
     "1. Seed the QA DB with the state described in the Pre-condition column.\n"
     "2. Load /{module}/dashboard as each role.\n"
     "3. Compare every rendered KPI value against the SQL / formula in the Verification Query column.\n"
     "4. Record result in Status. Do NOT fix failures — log them."),
    ("STATUS KEY",
     "Pass = UI value matches query | Fail = mismatch found | Pending = not yet run | N/A = role cannot see this KPI"),
]
for i, (k, v) in enumerate(rows):
    r = i + 4
    ws0.cell(r, 1, k).font = hfont(sz=9, color=H_BG)
    ws0.cell(r, 2, v).font = dfont()
    ws0.cell(r, 2).alignment = lft()
    ws0.row_dimensions[r].height = 48

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — MICRO LENDING DASHBOARD TESTS
# ══════════════════════════════════════════════════════════════════════════════
COLS = ["TC ID","Module","Role","KPI / Panel","Pre-condition (DB Seed State)","Verification Query / Formula","Expected Result","Status","Actual Value","Notes"]
WIDTHS = [10, 14, 12, 26, 48, 58, 38, 10, 18, 28]

def make_sheet(wb, title, tab_color):
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    ws.sheet_tab_color = tab_color
    ws.freeze_panes = "A2"
    for i, (col, w) in enumerate(zip(COLS, WIDTHS), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        ws.cell(1, i, col)
    hdr_row(ws, 1, len(COLS))
    return ws

def tc(ws, r, tc_id, module, role, kpi, precond, query, expected, alt=False):
    data_row(ws, r, [tc_id, module, role, kpi, precond, query, expected, "", "", ""], alt)

# ─── Micro Lending ────────────────────────────────────────────────────────────
ws1 = make_sheet(wb, "ML – Micro Lending", ML_HDR)
r = 2; alt = False

sec_row(ws1, r, "A. EXPECTED COLLECTION KPIs", len(COLS), ML_HDR); r += 1
tc(ws1, r, "ML-D-001","Micro Lending","admin/superadmin",
   "Today's Expected Collection",
   "2 active loans; 3 instalments due today (₹500, ₹300, ₹200). All status=upcoming.",
   "SELECT SUM(due_amount) FROM instalments\nWHERE loan.appType='microlending'\n  AND due_date >= today_start\n  AND due_date < tomorrow_start",
   "₹1,000 rendered on KPI card",alt); r+=1; alt=not alt

tc(ws1, r, "ML-D-002","Micro Lending","admin/superadmin",
   "Today's Expected — Zero state",
   "No instalments due today for any active loan.",
   "Same query → returns NULL/0",
   "₹0 rendered; no error or NaN",alt); r+=1; alt=not alt

sec_row(ws1, r, "B. TODAY'S COLLECTED KPIs", len(COLS), ML_HDR); r += 1
tc(ws1, r, "ML-D-003","Micro Lending","admin/superadmin",
   "Today's Collected (total)",
   "3 CollectionEntry rows submitted today: ₹500 cash, ₹300 UPI, ₹200 cash. All linked to microlending loans.",
   "SELECT SUM(received_amount) FROM collection_entries\nWHERE tenant_id=T\n  AND loan.appType='microlending'\n  AND submitted_at >= today_start\n  AND submitted_at < tomorrow_start",
   "₹1,000 rendered on KPI card",alt); r+=1; alt=not alt

tc(ws1, r, "ML-D-004","Micro Lending","admin/superadmin",
   "Today's Collected — cross-module isolation",
   "1 autofinance CollectionEntry ₹500 also submitted today.",
   "Same query scoped to appType='microlending' only",
   "₹1,000 — autofinance ₹500 must NOT be included",alt); r+=1; alt=not alt

tc(ws1, r, "ML-D-005","Micro Lending","admin/superadmin",
   "Today's Collected — cross-branch isolation",
   "Branch A has ₹1,000 collected. Branch B has ₹400 collected. Admin belongs to Branch A.",
   "Query filtered by branchId = Branch A",
   "₹1,000 — Branch B ₹400 must NOT appear",alt); r+=1; alt=not alt

sec_row(ws1, r, "C. OUTSTANDING ON TODAY'S DUES", len(COLS), ML_HDR); r += 1
tc(ws1, r, "ML-D-006","Micro Lending","admin/superadmin",
   "Outstanding on Today's Dues",
   "3 instalments due today: ₹500 (paid ₹200), ₹300 (unpaid), ₹200 (fully paid).",
   "SUM(GREATEST(0, due_amount - COALESCE(received_amount,0)))\nFOR instalments due today only",
   "₹400 (300 + 100 partial remainder)",alt); r+=1; alt=not alt

tc(ws1, r, "ML-D-007","Micro Lending","admin/superadmin",
   "Outstanding — all today's dues fully paid",
   "All today's instalments have received_amount >= due_amount.",
   "Same query → 0",
   "₹0 rendered on card",alt); r+=1; alt=not alt

tc(ws1, r, "ML-D-008","Micro Lending","admin/superadmin",
   "Outstanding — overpayment does NOT go negative",
   "One instalment due ₹300; received ₹400 (overpayment).",
   "GREATEST(0, 300-400) = 0 — must clamp",
   "₹0 (not -₹100); negative impossible",alt); r+=1; alt=not alt

sec_row(ws1, r, "D. ACTIVE CUSTOMERS", len(COLS), ML_HDR); r += 1
tc(ws1, r, "ML-D-009","Micro Lending","admin/superadmin",
   "Active Customers count",
   "5 customers: 3 status=active, 1 status=inactive, 1 status=pending_review.",
   "SELECT COUNT(*) FROM customers\nWHERE tenant_id=T AND appType='microlending'\n  AND status='active' AND branch_id=B",
   "3 rendered on KPI card",alt); r+=1; alt=not alt

tc(ws1, r, "ML-D-010","Micro Lending","admin/superadmin",
   "Active Customers — pending_review not counted",
   "1 customer status=pending_review.",
   "COUNT WHERE status='active' only",
   "Pending_review customer NOT counted",alt); r+=1; alt=not alt

sec_row(ws1, r, "E. OVERDUE KPIs", len(COLS), ML_HDR); r += 1
tc(ws1, r, "ML-D-011","Micro Lending","admin/superadmin",
   "Overdue Customers count",
   "4 instalments past due with outstanding > 0; owned by 3 distinct customers (one customer has 2 overdue instalments).",
   "SELECT COUNT(DISTINCT loan.customer_id)\nFROM instalments\nWHERE loan.appType='microlending'\n  AND due_date < today_start\n  AND status IN ('upcoming','missed','partial')\n  AND (due_amount - COALESCE(received_amount,0)) > 0",
   "3 (not 4 — deduplicated by customer)",alt); r+=1; alt=not alt

tc(ws1, r, "ML-D-012","Micro Lending","admin/superadmin",
   "Total Overdue Amount",
   "3 overdue instalments: outstanding ₹800, ₹600, ₹400.",
   "SUM(GREATEST(0, due_amount - received_amount))\nFOR all overdue instalments (no TAKE limit)",
   "₹1,800 — must use full set, not top-10 capped set",alt); r+=1; alt=not alt

tc(ws1, r, "ML-D-013","Micro Lending","admin/superadmin",
   "Overdue KPIs — partial payment reduces outstanding",
   "1 overdue instalment due ₹1,000; received ₹600.",
   "GREATEST(0, 1000-600) = 400",
   "₹400 shown (not ₹1,000 gross)",alt); r+=1; alt=not alt

tc(ws1, r, "ML-D-014","Micro Lending","admin/superadmin",
   "Overdue table — capped at 10 rows but KPI uses full set",
   "15 overdue instalments across 15 customers. Total outstanding = ₹15,000.",
   "KPI SUM uses unbounded query; table shows only top 10 by due_date asc",
   "KPI = ₹15,000; table shows 10 rows",alt); r+=1; alt=not alt

sec_row(ws1, r, "F. PENALTY ACCUMULATED", len(COLS), ML_HDR); r += 1
tc(ws1, r, "ML-D-015","Micro Lending","admin/superadmin",
   "Penalty Accumulated (pending+partial)",
   "3 penalties: gross ₹500 (settled ₹100, waived ₹0), ₹800 (settled ₹0, waived ₹200), ₹300 (settled ₹0, waived ₹0). All status=pending/partial.",
   "SUM(gross_penalty) - SUM(settled_amount) - SUM(waived_amount)\nWHERE status IN ('pending','partial')\n= (500+800+300)-(100+0+0)-(0+200+0) = 1300",
   "₹1,300 on KPI card",alt); r+=1; alt=not alt

tc(ws1, r, "ML-D-016","Micro Lending","admin/superadmin",
   "Penalty — settled penalties NOT included",
   "1 penalty status=settled, 1 status=pending.",
   "WHERE status IN ('pending','partial') only",
   "Settled penalty excluded from total",alt); r+=1; alt=not alt

tc(ws1, r, "ML-D-017","Micro Lending","admin/superadmin",
   "Penalty — result never goes negative",
   "grossPenalty=100, settledAmount=60, waivedAmount=60 (sum exceeds gross).",
   "GREATEST(0, gross-settled-waived) = GREATEST(0,100-120) = 0",
   "₹0 rendered (not -₹20)",alt); r+=1; alt=not alt

sec_row(ws1, r, "G. PENDING APPROVALS", len(COLS), ML_HDR); r += 1
tc(ws1, r, "ML-D-018","Micro Lending","admin/superadmin",
   "Pending Approvals count",
   "2 ApprovalRequest rows status=pending; 1 loan status=pending_review; 1 customer status=pending_review.",
   "COUNT(approval_requests WHERE status='pending')\n+ COUNT(loans WHERE status='pending_review')\n+ COUNT(customers WHERE status='pending_review')\n= 2+1+1 = 4",
   "4 on KPI badge",alt); r+=1; alt=not alt

tc(ws1, r, "ML-D-019","Micro Lending","admin/superadmin",
   "Pending Approvals — approved requests not counted",
   "2 ApprovalRequest rows status=approved; 0 pending.",
   "COUNT = 0",
   "0 on badge; no badge shown or badge shows 0",alt); r+=1; alt=not alt

sec_row(ws1, r, "H. CAPITAL BALANCE", len(COLS), ML_HDR); r += 1
tc(ws1, r, "ML-D-020","Micro Lending","admin/superadmin",
   "Capital Balance formula",
   "AccountEntries: capital_add ₹100,000; loan_disburse ₹80,000; collection ₹85,000; expense ₹5,000.",
   "100000 - 80000 + 85000 - 5000 = 100,000",
   "₹1,00,000 on KPI card",alt); r+=1; alt=not alt

tc(ws1, r, "ML-D-021","Micro Lending","admin/superadmin",
   "Capital Balance — negative allowed",
   "capital_add ₹10,000; loan_disburse ₹15,000.",
   "10000 - 15000 = -5000",
   "-₹5,000 shown with red icon",alt); r+=1; alt=not alt

tc(ws1, r, "ML-D-022","Micro Lending","admin/superadmin",
   "Pending Field Float banner",
   "2 pending UPI entries ₹500+₹300; 1 pending cash entry ₹200.",
   "pendingUPI(800) + pendingCash(200) = 1000",
   "'+ ₹1,000 pending field float' shown below capital",alt); r+=1; alt=not alt

sec_row(ws1, r, "I. TODAY'S COLLECTION SPLIT (Cash vs UPI)", len(COLS), ML_HDR); r += 1
tc(ws1, r, "ML-D-023","Micro Lending","admin/superadmin",
   "Cash total in split panel",
   "Today: 3 cash entries ₹200, ₹300, ₹500.",
   "SUM(received_amount) WHERE payment_mode='cash' AND submitted_at=today",
   "Cash: ₹1,000",alt); r+=1; alt=not alt

tc(ws1, r, "ML-D-024","Micro Lending","admin/superadmin",
   "UPI total in split panel",
   "Today: 2 UPI entries ₹400, ₹600; 1 online entry ₹200.",
   "SUM WHERE payment_mode IN ('upi','online') AND submitted_at=today",
   "UPI: ₹1,200",alt); r+=1; alt=not alt

tc(ws1, r, "ML-D-025","Micro Lending","admin/superadmin",
   "Split bar proportions",
   "Cash=₹1,000, UPI=₹1,200 (total ₹2,200).",
   "cash_width% = 1000/2200*100 = 45.5%; upi_width% = 54.5%",
   "Bar widths in correct proportion; total label shows ₹2,200",alt); r+=1; alt=not alt

tc(ws1, r, "ML-D-026","Micro Lending","admin/superadmin",
   "Split — zero collections today",
   "No CollectionEntry submitted today.",
   "Both sums = 0; bar shows empty grey strip",
   "No crash; grey bar shown; no NaN",alt); r+=1; alt=not alt

sec_row(ws1, r, "J. COLLECTION DETAILS TABLE (Route-wise)", len(COLS), ML_HDR); r += 1
tc(ws1, r, "ML-D-027","Micro Lending","admin/superadmin",
   "Collected Today per route",
   "Route A: 2 collection entries today ₹400+₹600. Route B: 1 entry ₹300.",
   "SUM(received_amount) WHERE customer.route_id = route.id AND submitted_at=today",
   "Route A = ₹1,000; Route B = ₹300",alt); r+=1; alt=not alt

tc(ws1, r, "ML-D-028","Micro Lending","admin/superadmin",
   "Overdue per route in table",
   "Route A customers have 3 overdue instalments outstanding ₹200+₹300+₹500.",
   "SUM outstanding on past-due instalments for customers on this route",
   "Route A overdue = ₹1,000",alt); r+=1; alt=not alt

tc(ws1, r, "ML-D-029","Micro Lending","admin/superadmin",
   "Collect Cash button appears when pending cash",
   "Agent has submitted ₹500 cash for Route A; verificationStatus=pending.",
   "pendingCashCollections filter by routeId",
   "Collect Cash button visible with ₹500 amount for Route A",alt); r+=1; alt=not alt

tc(ws1, r, "ML-D-030","Micro Lending","admin/superadmin",
   "No routes — empty state shown",
   "No Route rows in DB for this tenant/branch.",
   "routes query returns []",
   "Empty state card shown; no JS error",alt); r+=1; alt=not alt

sec_row(ws1, r, "K. OVERDUE ALERTS TABLE", len(COLS), ML_HDR); r += 1
tc(ws1, r, "ML-D-031","Micro Lending","admin/superadmin",
   "Overdue table — correct customer name and days",
   "Instalment due 5 days ago; outstanding ₹700; customer='Ravi Kumar'.",
   "daysOverdue = today - dueDate = 5",
   "Row shows 'Ravi Kumar', '₹700', '5d overdue'",alt); r+=1; alt=not alt

tc(ws1, r, "ML-D-032","Micro Lending","admin/superadmin",
   "Overdue table — ordered oldest first",
   "3 overdue instalments: 1d, 10d, 5d overdue.",
   "ORDER BY due_date ASC",
   "Row order: 10d first, then 5d, then 1d",alt); r+=1; alt=not alt

sec_row(ws1, r, "L. PENDING UPI VERIFICATIONS", len(COLS), ML_HDR); r += 1
tc(ws1, r, "ML-D-033","Micro Lending","admin/superadmin",
   "Pending UPI — count and total",
   "3 CollectionEntry rows: paymentMode=upi, verificationStatus=pending, amounts ₹200+₹400+₹600.",
   "SELECT * FROM collection_entries WHERE payment_mode IN ('upi','online') AND verification_status='pending'",
   "3 rows in table; Bulk Verify button shows ₹1,200",alt); r+=1; alt=not alt

tc(ws1, r, "ML-D-034","Micro Lending","admin/superadmin",
   "Pending UPI — none present",
   "No unverified UPI entries.",
   "Query returns []",
   "Empty state shown with checkmark; no table rendered",alt); r+=1; alt=not alt

tc(ws1, r, "ML-D-035","Micro Lending","admin/superadmin",
   "Pending UPI — cash entries NOT shown",
   "1 cash entry verificationStatus=pending; 1 UPI entry verificationStatus=pending.",
   "Filter: payment_mode IN ('upi','online') only",
   "Only UPI entry shown; cash entry excluded",alt); r+=1; alt=not alt

sec_row(ws1, r, "M. 7-DAY COLLECTION TREND", len(COLS), ML_HDR); r += 1
tc(ws1, r, "ML-D-036","Micro Lending","admin/superadmin",
   "Trend — 7 bars rendered",
   "Any state with data in the last 7 days.",
   "Array.from({ length: 7 }) — always 7 data points",
   "Exactly 7 bars visible in chart; labels are DD-MMM format",alt); r+=1; alt=not alt

tc(ws1, r, "ML-D-037","Micro Lending","admin/superadmin",
   "Trend — collected bar capped at due amount",
   "Day 3: dueAmount=₹1,000; received=₹1,500 (overpayment).",
   "collected = MIN(received_amount, due_amount) = 1000",
   "Collected bar height = 100%; not exceeding expected bar",alt); r+=1; alt=not alt

tc(ws1, r, "ML-D-038","Micro Lending","admin/superadmin",
   "Trend — day with zero dues shows zero bars",
   "Sunday: no instalments due.",
   "expected=0, collected=0",
   "Both bars render at minimum height (4px); no NaN/Inf",alt); r+=1; alt=not alt

sec_row(ws1, r, "N. RECENT LOANS THIS MONTH", len(COLS), ML_HDR); r += 1
tc(ws1, r, "ML-D-039","Micro Lending","admin/superadmin",
   "Recent Loans count (this month)",
   "5 loans created this calendar month; 2 created last month.",
   "SELECT COUNT(*) FROM loans WHERE created_at >= first_of_month",
   "5 on KPI card (last month's 2 excluded)",alt); r+=1; alt=not alt

sec_row(ws1, r, "O. BEST PAYER & HIGHEST BORROWER", len(COLS), ML_HDR); r += 1
tc(ws1, r, "ML-D-040","Micro Lending","admin/superadmin",
   "Highest Active Borrower",
   "Customer A: 1 active loan ₹50,000. Customer B: 2 active loans ₹30,000+₹40,000 = ₹70,000.",
   "GROUP BY customer_id, SUM(principal) ORDER BY SUM DESC LIMIT 1",
   "Customer B shown as Highest Borrower",alt); r+=1; alt=not alt

tc(ws1, r, "ML-D-041","Micro Lending","admin/superadmin",
   "Best Payer (zero overdue, paid count > 0)",
   "Customer A: paidCount=5, no missed instalments. Customer B: paidCount=0.",
   "WHERE paidCount > 0 AND instalments.none(status='missed')",
   "Customer A shown; Customer B excluded",alt); r+=1; alt=not alt


# ─── Auto Finance ─────────────────────────────────────────────────────────────
ws2 = make_sheet(wb, "AF – Auto Finance", AF_HDR)
r = 2; alt = False

sec_row(ws2, r, "A. EXPECTED & COLLECTED KPIs — SHARED WITH ML (SCOPED TO autofinance)", len(COLS), AF_HDR); r += 1
tc(ws2, r, "AF-D-001","Auto Finance","admin/superadmin",
   "Today's Expected — auto finance scope",
   "3 instalments due today belonging to autofinance loans (₹1,000, ₹800, ₹700). ML loans also exist.",
   "WHERE loan.appType='autofinance' AND due_date=today",
   "₹2,500 — ML instalments NOT included",alt); r+=1; alt=not alt

tc(ws2, r, "AF-D-002","Auto Finance","admin/superadmin",
   "Today's Collected — auto finance scope",
   "CollectionEntries today: 2 from autofinance loans (₹600, ₹900); 1 from ML loan (₹300).",
   "WHERE loan.appType='autofinance' AND submitted_at=today",
   "₹1,500 — ML ₹300 NOT included",alt); r+=1; alt=not alt

tc(ws2, r, "AF-D-003","Auto Finance","admin/superadmin",
   "Outstanding on Today's Dues — auto finance",
   "2 AF instalments due today: ₹1,000 (paid ₹400), ₹800 (unpaid).",
   "SUM outstanding = 600 + 800 = 1400",
   "₹1,400",alt); r+=1; alt=not alt

tc(ws2, r, "AF-D-004","Auto Finance","admin/superadmin",
   "Active Customers — auto finance",
   "4 customers appType=autofinance, 3 active + 1 inactive.",
   "WHERE appType='autofinance' AND status='active'",
   "3",alt); r+=1; alt=not alt

tc(ws2, r, "AF-D-005","Auto Finance","admin/superadmin",
   "Overdue Amount — auto finance",
   "2 AF overdue instalments outstanding ₹500+₹700. 1 ML overdue ₹400.",
   "WHERE loan.appType='autofinance' AND overdue condition",
   "₹1,200 — ML ₹400 excluded",alt); r+=1; alt=not alt

tc(ws2, r, "AF-D-006","Auto Finance","admin/superadmin",
   "Penalty Accumulated — auto finance",
   "2 AF penalties pending: gross ₹600 settled ₹100, gross ₹400 settled ₹0.",
   "SUM(gross-settled-waived) for AF penalties",
   "₹900",alt); r+=1; alt=not alt

tc(ws2, r, "AF-D-007","Auto Finance","admin/superadmin",
   "Pending Approvals — auto finance",
   "1 ApprovalRequest appType=autofinance status=pending; 1 appType=microlending status=pending.",
   "WHERE appType='autofinance' AND status='pending'",
   "1 — ML request excluded",alt); r+=1; alt=not alt

sec_row(ws2, r, "B. AUTO FINANCE — VEHICLE-LINKED LOAN DATA", len(COLS), AF_HDR); r += 1
tc(ws2, r, "AF-D-008","Auto Finance","admin/superadmin",
   "Dashboard shows AF loans (with vehicles) in collection details",
   "Route X has 2 AF customers with vehicle loans. Collection today ₹1,200.",
   "Collection detail table scoped to appType='autofinance'",
   "Route X shows ₹1,200 collected; no ML loans bleed in",alt); r+=1; alt=not alt

tc(ws2, r, "AF-D-009","Auto Finance","admin/superadmin",
   "Overdue alert — vehicle loan customer",
   "Customer with vehicle loan; instalment 7 days overdue ₹900.",
   "Overdue table query scoped to appType='autofinance'",
   "Customer shown with ₹900, '7d overdue'",alt); r+=1; alt=not alt

sec_row(ws2, r, "C. COLLECTION SPLIT (CASH vs UPI) — AF SCOPE", len(COLS), AF_HDR); r += 1
tc(ws2, r, "AF-D-010","Auto Finance","admin/superadmin",
   "Cash/UPI split — scoped to AF loans only",
   "Cash ₹800 from AF loan; Cash ₹500 from ML loan. UPI ₹400 from AF loan.",
   "WHERE loan.appType='autofinance' AND submitted_at=today",
   "Cash: ₹800, UPI: ₹400 (ML ₹500 excluded)",alt); r+=1; alt=not alt

sec_row(ws2, r, "D. 7-DAY TREND — AF", len(COLS), AF_HDR); r += 1
tc(ws2, r, "AF-D-011","Auto Finance","admin/superadmin",
   "Trend bars — AF data only",
   "Day 1: AF instalments due ₹2,000 collected ₹1,500. ML instalments due ₹3,000.",
   "weekInstalments WHERE loan.appType='autofinance'",
   "Day 1 expected bar = ₹2,000; ML ₹3,000 not in chart",alt); r+=1; alt=not alt

sec_row(ws2, r, "E. PENDING UPI — AF", len(COLS), AF_HDR); r += 1
tc(ws2, r, "AF-D-012","Auto Finance","admin/superadmin",
   "Pending UPI — AF scoped",
   "1 UPI pending from AF loan; 1 UPI pending from ML loan.",
   "WHERE loan.appType='autofinance'",
   "Only 1 AF UPI entry shown; ML excluded",alt); r+=1; alt=not alt

sec_row(ws2, r, "F. CAPITAL BALANCE — AF (shared accounting entries)", len(COLS), AF_HDR); r += 1
tc(ws2, r, "AF-D-013","Auto Finance","admin/superadmin",
   "Capital Balance — branch-scoped",
   "Branch B: capital_add ₹50,000; loan_disburse ₹30,000; collection ₹32,000. Branch A has separate entries.",
   "AccountEntry WHERE branch_id = Branch B",
   "₹52,000 (Branch A entries excluded)",alt); r+=1; alt=not alt


# ─── Chit Funds ───────────────────────────────────────────────────────────────
ws3 = make_sheet(wb, "CF – Chit Funds", CF_HDR)
r = 2; alt = False

sec_row(ws3, r, "A. CHIT FUND KPI CARDS", len(COLS), CF_HDR); r += 1
tc(ws3, r, "CF-D-001","Chit Funds","admin/superadmin",
   "Active Chit Groups count",
   "5 ChitGroup rows: 3 status=active, 1 status=closed, 1 status=draft.",
   "SELECT COUNT(*) FROM chit_groups WHERE status='active' AND appType='chitfunds'",
   "3 on KPI card",alt); r+=1; alt=not alt

tc(ws3, r, "CF-D-002","Chit Funds","admin/superadmin",
   "Total Subscribers count",
   "Group A: 10 members; Group B: 8 members; Group C (closed): 5 members.",
   "SELECT COUNT(*) FROM chit_members WHERE chit_group.tenant_id=T\n(ALL members, regardless of group status)",
   "23 (all members including closed group)",alt); r+=1; alt=not alt

tc(ws3, r, "CF-D-003","Chit Funds","admin/superadmin",
   "Auctions This Month count",
   "3 auctions this calendar month; 2 from last month.",
   "WHERE auction_date >= first_of_month AND auction_date < tomorrow",
   "3 on KPI card",alt); r+=1; alt=not alt

tc(ws3, r, "CF-D-004","Chit Funds","admin/superadmin",
   "Pending Approvals (CF scope)",
   "2 ApprovalRequest appType=chitfunds status=pending; 1 appType=microlending.",
   "WHERE appType='chitfunds' AND status='pending'",
   "2 on badge; ML request excluded",alt); r+=1; alt=not alt

sec_row(ws3, r, "B. TODAY'S SUBSCRIPTION KPIs", len(COLS), CF_HDR); r += 1
tc(ws3, r, "CF-D-005","Chit Funds","admin/superadmin",
   "Today's Expected Subscriptions",
   "4 ChitSubscription rows due today: ₹500, ₹500, ₹1,000, ₹750.",
   "SELECT SUM(due_amount) FROM chit_subscriptions\nWHERE due_date >= today AND due_date < tomorrow",
   "₹2,750",alt); r+=1; alt=not alt

tc(ws3, r, "CF-D-006","Chit Funds","admin/superadmin",
   "Subscriptions Collected Today",
   "Same 4 subscriptions: paid ₹500, ₹0, ₹800, ₹750.",
   "SELECT SUM(paid_amount) for today's subscriptions",
   "₹2,050",alt); r+=1; alt=not alt

tc(ws3, r, "CF-D-007","Chit Funds","admin/superadmin",
   "Today's Balance Due",
   "Expected ₹2,750; Collected ₹2,050.",
   "GREATEST(0, expected - collected) = 700",
   "₹700",alt); r+=1; alt=not alt

tc(ws3, r, "CF-D-008","Chit Funds","admin/superadmin",
   "Balance Due never negative",
   "Expected ₹1,000; Paid ₹1,200 (overpayment).",
   "GREATEST(0, 1000-1200) = 0",
   "₹0 (not -₹200)",alt); r+=1; alt=not alt

sec_row(ws3, r, "C. OVERDUE CHIT SUBSCRIPTIONS", len(COLS), CF_HDR); r += 1
tc(ws3, r, "CF-D-009","Chit Funds","admin/superadmin",
   "Total Overdue Contributions amount",
   "3 ChitSubscription rows past due: outstanding ₹500, ₹750, ₹1,000.",
   "SUM(GREATEST(0, due_amount - paid_amount)) WHERE due_date < today AND status != 'paid'",
   "₹2,250",alt); r+=1; alt=not alt

tc(ws3, r, "CF-D-010","Chit Funds","admin/superadmin",
   "Overdue Members count (deduplicated)",
   "3 overdue subscriptions: 2 belong to same member.",
   "COUNT(DISTINCT member_id) for overdue subscriptions",
   "2 (not 3)",alt); r+=1; alt=not alt

tc(ws3, r, "CF-D-011","Chit Funds","admin/superadmin",
   "Overdue table — max 10 rows shown",
   "15 overdue subscriptions.",
   "Query has TAKE 10",
   "Table shows 10 rows; total KPI uses full set",alt); r+=1; alt=not alt

tc(ws3, r, "CF-D-012","Chit Funds","admin/superadmin",
   "Overdue table — correct days overdue",
   "Subscription due 8 days ago.",
   "FLOOR((today - due_date) / 86400000) = 8",
   "Row shows '8d' in days column",alt); r+=1; alt=not alt

tc(ws3, r, "CF-D-013","Chit Funds","admin/superadmin",
   "Overdue — zero outstanding: not shown",
   "1 subscription past due date but paid_amount = due_amount.",
   "outstanding = 0 → excluded from overdue set",
   "This subscription does NOT appear in table",alt); r+=1; alt=not alt

sec_row(ws3, r, "D. ACTIVE CHIT GROUPS TABLE", len(COLS), CF_HDR); r += 1
tc(ws3, r, "CF-D-014","Chit Funds","admin/superadmin",
   "Active Chit Groups list — max 5 shown",
   "7 active chit groups.",
   "Query has TAKE 5",
   "Table shows 5 rows only",alt); r+=1; alt=not alt

tc(ws3, r, "CF-D-015","Chit Funds","admin/superadmin",
   "Members column: enrolled / total",
   "Group X: totalMembers=20, actual ChitMember rows=15.",
   "_count.members=15, group.totalMembers=20",
   "Shows '15 / 20'",alt); r+=1; alt=not alt

tc(ws3, r, "CF-D-016","Chit Funds","admin/superadmin",
   "Chit Value shown correctly",
   "Group Y: chitValue=50000.",
   "formatCurrency(50000)",
   "₹50,000 rendered",alt); r+=1; alt=not alt

tc(ws3, r, "CF-D-017","Chit Funds","admin/superadmin",
   "Empty groups — empty state shown",
   "No active ChitGroup rows.",
   "activeChitGroupsList = []",
   "Empty state row 'No active chit groups found.' shown",alt); r+=1; alt=not alt

sec_row(ws3, r, "E. CF MODULE ISOLATION", len(COLS), CF_HDR); r += 1
tc(ws3, r, "CF-D-018","Chit Funds","admin/superadmin",
   "CF dashboard does not include ML or AF data",
   "ML: 10 active customers; AF: 5 active customers; CF: 3 active members.",
   "All CF queries scoped to appType='chitfunds'",
   "KPI shows CF data only; ML/AF data invisible",alt); r+=1; alt=not alt

tc(ws3, r, "CF-D-019","Chit Funds","admin/superadmin",
   "CF redirect — /chitfunds/dashboard redirects to /chitfunds/chits",
   "appType resolved as 'chitfunds'.",
   "if (appType === 'chitfunds') redirect('/chits')",
   "Browser lands on /chitfunds/chits, not dashboard",alt); r+=1; alt=not alt


# ─── Agent Dashboard ──────────────────────────────────────────────────────────
ws4 = make_sheet(wb, "AG – Agent View (All Modules)", AG_HDR)
r = 2; alt = False

sec_row(ws4, r, "A. AGENT KPI CARDS — SCOPED TO ASSIGNED ROUTES ONLY", len(COLS), AG_HDR); r += 1
tc(ws4, r, "AG-D-001","Micro Lending","agent",
   "My Expected Collection — route-scoped",
   "Agent assigned to Route A (3 customers). Route B exists (different agent). Route A instalments due today: ₹300+₹400.",
   "WHERE loan.customer.route.routeAgents.some(agentId) AND dueDate=today",
   "₹700 — Route B instalments excluded",alt); r+=1; alt=not alt

tc(ws4, r, "AG-D-002","Micro Lending","agent",
   "My Collected Today — agent's own entries only",
   "Agent A collected ₹600 today. Agent B collected ₹400 today (different route).",
   "WHERE agentId = agent.id AND submitted_at=today",
   "₹600 — Agent B's ₹400 not shown",alt); r+=1; alt=not alt

tc(ws4, r, "AG-D-003","Micro Lending","agent",
   "My Remaining Balance",
   "Expected ₹700, collected ₹600.",
   "GREATEST(0, 700-600) = 100",
   "₹100",alt); r+=1; alt=not alt

tc(ws4, r, "AG-D-004","Micro Lending","agent",
   "My Active Customers count",
   "Route A: 3 active customers. Route B: 4 active customers (not this agent's).",
   "WHERE route.routeAgents.some(agentId=me) AND status='active'",
   "3 — Route B customers excluded",alt); r+=1; alt=not alt

sec_row(ws4, r, "B. PROGRESS BAR", len(COLS), AG_HDR); r += 1
tc(ws4, r, "AG-D-005","Micro Lending","agent",
   "Progress bar percentage",
   "Expected ₹1,000; Collected ₹750.",
   "Math.round(750/1000*100) = 75%",
   "Progress bar renders at 75%; label shows '75% Completed'",alt); r+=1; alt=not alt

tc(ws4, r, "AG-D-006","Micro Lending","agent",
   "Progress bar — 100% when expected = 0",
   "No instalments due today for agent's routes.",
   "expected=0 → progressPct=100",
   "100% shown; no division by zero error",alt); r+=1; alt=not alt

tc(ws4, r, "AG-D-007","Micro Lending","agent",
   "Progress bar — capped at 100% even if over-collected",
   "Expected ₹500; Collected ₹700.",
   "MIN(100, 700/500*100) = 100",
   "Bar shows 100%; does not overflow",alt); r+=1; alt=not alt

sec_row(ws4, r, "C. ASSIGNED ROUTES TABLE", len(COLS), AG_HDR); r += 1
tc(ws4, r, "AG-D-008","Micro Lending","agent",
   "Only agent's routes shown",
   "Agent assigned to Route A and Route C. Route B and D exist for other agents.",
   "WHERE routeAgents.some(agentId=me)",
   "Only Route A and Route C rows in table",alt); r+=1; alt=not alt

tc(ws4, r, "AG-D-009","Micro Lending","agent",
   "Route overdue amount — scoped to route's customers",
   "Route A customers have 2 overdue instalments outstanding ₹300+₹600.",
   "Sum outstanding for past-due instalments on Route A customers",
   "₹900 in Overdue column for Route A",alt); r+=1; alt=not alt

tc(ws4, r, "AG-D-010","Micro Lending","agent",
   "Zero overdue — green colour shown",
   "Route A has no overdue instalments.",
   "overdue = 0 → color=success",
   "₹0 shown in green",alt); r+=1; alt=not alt

sec_row(ws4, r, "D. MY RECENT APPROVAL REQUESTS", len(COLS), AG_HDR); r += 1
tc(ws4, r, "AG-D-011","Micro Lending","agent",
   "Shows only this agent's requests",
   "Agent A has 3 ApprovalRequest rows. Agent B has 2. All appType=microlending.",
   "WHERE requestedById = agent.id ORDER BY createdAt DESC LIMIT 5",
   "3 rows shown — Agent B's 2 excluded",alt); r+=1; alt=not alt

tc(ws4, r, "AG-D-012","Micro Lending","agent",
   "Pending request shows 'Waiting for Admin approval'",
   "1 pending request with no reviewNotes, no reviewedBy.",
   "reviewNotes = null AND reviewedBy = null",
   "'Waiting for Admin approval...' in Notes column",alt); r+=1; alt=not alt

tc(ws4, r, "AG-D-013","Micro Lending","agent",
   "Approved request shows reviewer note",
   "1 approved request: reviewNotes='Verified by admin', reviewedBy=Admin A.",
   "reviewNotes present → show notes",
   "'Verified by admin' shown in Notes column",alt); r+=1; alt=not alt

sec_row(ws4, r, "E. MY OVERDUE ALERTS", len(COLS), AG_HDR); r += 1
tc(ws4, r, "AG-D-014","Micro Lending","agent",
   "Only agent's customers shown in overdue alerts",
   "Route A (this agent): 2 overdue customers. Route B (other agent): 1 overdue customer.",
   "WHERE loan.customer.route.routeAgents.some(agentId=me)",
   "2 rows — Route B customer excluded",alt); r+=1; alt=not alt

tc(ws4, r, "AG-D-015","Micro Lending","agent",
   "Max 5 overdue alerts shown",
   "Agent has 8 overdue customers.",
   "overdueInstalments.slice(0,5)",
   "5 rows shown; remaining 3 not visible",alt); r+=1; alt=not alt

tc(ws4, r, "AG-D-016","Micro Lending","agent",
   "Zero overdue — success state shown",
   "Agent has no overdue instalments on assigned routes.",
   "overdueInstalments = [] → show checkmark",
   "'All collections up to date!' shown in green",alt); r+=1; alt=not alt

sec_row(ws4, r, "F. AGENT VIEW — ROLE GATES (KPIs NOT shown to agent)", len(COLS), AG_HDR); r += 1
tc(ws4, r, "AG-D-017","Micro Lending","agent",
   "Capital Balance NOT shown to agent",
   "Agent views /{module}/dashboard.",
   "getAgentDashboardData() — no currentCapital returned",
   "No Capital Balance card rendered in agent dashboard",alt); r+=1; alt=not alt

tc(ws4, r, "AG-D-018","Micro Lending","agent",
   "Cash/UPI split NOT shown to agent",
   "Agent views dashboard.",
   "Agent dashboard section has no split panel",
   "No 'Today's Collection Split' bar rendered",alt); r+=1; alt=not alt

tc(ws4, r, "AG-D-019","Micro Lending","agent",
   "Pending UPI panel NOT shown to agent",
   "Agent views dashboard.",
   "No pendingUpiCollections in agent data",
   "No Pending UPI Verifications table rendered",alt); r+=1; alt=not alt

tc(ws4, r, "AG-D-020","Micro Lending","agent",
   "Collection Details (route table) NOT shown to agent",
   "Agent views dashboard.",
   "Agent gets routePerformance list (no agent column, no collected-today column)",
   "Only Assigned Routes table shown (simplified); no Collect Cash button",alt); r+=1; alt=not alt

sec_row(ws4, r, "G. AGENT DASHBOARD 7-DAY TREND", len(COLS), AG_HDR); r += 1
tc(ws4, r, "AG-D-021","Micro Lending","agent",
   "Trend chart scoped to agent's loans only",
   "Agent's route: 7-day instalments expected ₹5,000. Other routes: ₹8,000.",
   "weekInstalments WHERE loan.customer.route.routeAgents.some(agentId=me)",
   "Trend chart reflects only ₹5,000 total; not ₹13,000",alt); r+=1; alt=not alt


# ─── Cross-Module & Edge Cases ────────────────────────────────────────────────
ws5 = make_sheet(wb, "XM – Cross-Module & Edge Cases", "7030A0")
r = 2; alt = False

sec_row(ws5, r, "A. MODULE URL ROUTING", len(COLS), "7030A0"); r += 1
tc(ws5, r, "XM-D-001","All","admin/superadmin",
   "/microlending/dashboard loads ML data",
   "Session active_app_type=autofinance; URL = /microlending/dashboard.",
   "getUserAppType() reads URL segment first → returns 'microlending'",
   "ML KPIs shown; AF data not shown",alt); r+=1; alt=not alt

tc(ws5, r, "XM-D-002","All","admin/superadmin",
   "/autofinance/dashboard loads AF data",
   "Same session; URL = /autofinance/dashboard.",
   "getUserAppType() → 'autofinance'",
   "AF KPIs shown; ML data not shown",alt); r+=1; alt=not alt

tc(ws5, r, "XM-D-003","All","admin/superadmin",
   "/chitfunds/dashboard redirects to /chitfunds/chits",
   "URL = /chitfunds/dashboard.",
   "redirect('/chits') triggered",
   "Browser lands on /chitfunds/chits",alt); r+=1; alt=not alt

tc(ws5, r, "XM-D-004","All","superadmin",
   "Branch switcher — KPIs update for new branch",
   "Branch A: 5 active customers. Branch B: 12 active customers. Superadmin switches to Branch B.",
   "active_branch_id cookie = Branch B; getDashboardData filters by branchId=Branch B",
   "Active Customers shows 12 after switch",alt); r+=1; alt=not alt

tc(ws5, r, "XM-D-005","All","admin/superadmin",
   "Penalty sync runs on every dashboard load",
   "2 instalments past due; no penalty rows yet.",
   "ensurePendingPenaltiesForMissedLoans called; 2 penalties created",
   "Penalty Accumulated card shows correct total after first load",alt); r+=1; alt=not alt

sec_row(ws5, r, "B. CONCURRENT / REAL-TIME UPDATES", len(COLS), "7030A0"); r += 1
tc(ws5, r, "XM-D-006","All","admin/superadmin",
   "Collection submitted → Dashboard reflects on next load",
   "Agent submits ₹500 collection; admin reloads /microlending/dashboard.",
   "todayCollected re-queried on each page load (no cache)",
   "₹500 appears in Today's Collected immediately",alt); r+=1; alt=not alt

tc(ws5, r, "XM-D-007","All","admin/superadmin",
   "New loan disbursed → Capital decreases on next load",
   "loan_disburse AccountEntry ₹20,000 added.",
   "currentCapital recalculated from all entries on each load",
   "Capital card decreases by ₹20,000",alt); r+=1; alt=not alt

tc(ws5, r, "XM-D-008","All","admin/superadmin",
   "Penalty waived → Penalty Accumulated decreases",
   "Penalty waivedAmount updated in DB.",
   "pendingPenalties re-aggregated on dashboard load",
   "Penalty card reflects waive",alt); r+=1; alt=not alt

sec_row(ws5, r, "C. EMPTY DATABASE STATES", len(COLS), "7030A0"); r += 1
tc(ws5, r, "XM-D-009","All","admin/superadmin",
   "Fresh tenant — all KPIs show zero",
   "No customers, loans, penalties, accounting entries for this tenant.",
   "All aggregate queries return NULL/0",
   "All KPI cards show ₹0 or 0; no JS errors, no NaN, no blank cards",alt); r+=1; alt=not alt

tc(ws5, r, "XM-D-010","Chit Funds","admin/superadmin",
   "Fresh CF tenant — zero active groups",
   "No ChitGroup rows.",
   "totalChitGroups = 0",
   "KPI shows 0; empty state in groups table",alt); r+=1; alt=not alt

sec_row(ws5, r, "D. FORMATTING & DISPLAY", len(COLS), "7030A0"); r += 1
tc(ws5, r, "XM-D-011","All","admin/superadmin",
   "Indian currency format — 7 digit amount",
   "currentCapital = 9999999",
   "formatCurrency(9999999)",
   "₹99,99,999 (Indian grouping, not ₹9,999,999)",alt); r+=1; alt=not alt

tc(ws5, r, "XM-D-012","All","admin/superadmin",
   "Date labels in trend use DD-MMM format",
   "Any date in trend array.",
   "toLocaleDateString('en-IN', {day:'2-digit', month:'short'})",
   "Labels show e.g. '15 May' not '5/15/2026'",alt); r+=1; alt=not alt

tc(ws5, r, "XM-D-013","All","superadmin",
   "Recent Activity — developer actions excluded",
   "4 audit logs: 2 from admin, 1 from agent, 1 from developer.",
   "WHERE user.role != 'developer'",
   "3 activity items shown; developer action excluded",alt); r+=1; alt=not alt

tc(ws5, r, "XM-D-014","All","admin/superadmin",
   "Recent Activity — max 8 items",
   "15 audit log entries.",
   "TAKE 8",
   "8 items in activity feed; oldest 7 not shown",alt); r+=1; alt=not alt


# ── Save ──────────────────────────────────────────────────────────────────────
import os
output_dir = os.path.dirname(__file__)
out = os.path.join(output_dir, "LoanTrack_Dashboard_KPI_Test_Plan.xlsx")
wb.save(out)
print(f"Saved: {out}")

# Count all TCs
total = 0
for ws in [ws1, ws2, ws3, ws4, ws5]:
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and str(row[0]).count('-D-'):
            total += 1
print(f"Total test cases: {total}")
